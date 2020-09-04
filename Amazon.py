from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

class Amazon:

  def __init__(self):
    options = webdriver.ChromeOptions()
    options.add_argument('lang=pt-br')
    self.driver =  webdriver.Chrome(executable_path=r'./chromedriver')
    self.title = []
    self.price = []
  
  def PriceList(self):
    self.driver.get('https://www.amazon.com.br/')
    textbox = self.driver.find_element_by_id('twotabsearchtextbox')
    textbox.click()
    textbox.send_keys('iphone')
    search = self.driver.find_element_by_xpath('//*[@id="nav-search"]/form/div[2]/div')
    search.click()    
    
    time.sleep(5)
    data = self.driver.find_element_by_id('a-page')
    html = data.get_attribute("innerHTML")
    soup = BeautifulSoup(html, 'html.parser')
    
    for link in soup.findAll("div", {"class": "s-expand-height s-include-content-margin s-border-bottom s-latency-cf-section"}):
      res = link.get_text()
      #print(res)
      self.title.append(res)
      res.split('\n')
      print (self.title)

    wb = Workbook()
    ws = wb.active

    for i in range(len(self.title)):
      ws['A' + str(i + 1)] = self.title[i]
     
    wb.save('./Amazon.xlsx')   
    self.driver.close()
     
buy = Amazon()
buy.PriceList()
    


  # Primeiro teste que deu quase certo, porem como alguns protudos vinham sem precço, a função quebrava 


    # for link in soup.findAll("span", {"class": "a-size-base-plus a-color-base a-text-normal"}):
    #   res = link.get_text()
    #   self.title.append(res)
    
    # for link2 in soup.findAll("span", {"class": "a-offscreen"}):
    #   res2 = link2.get_text()
    #   self.price.append(res2)

    # teste = len(self.title)
    # teste2 = len(self.price)
    # print(teste, teste2)

    # for link in soup.findAll("span", {"class": "a-offscreen"}):
    #   res2 = link.get_text()
    #   self.price.append(res2) 

    # for i in range(len(self.title)):
    #   ws['A' + str(i + 1)] = self.title[i]
    #   ws['B' + str(i + 1)] = self.price[i]